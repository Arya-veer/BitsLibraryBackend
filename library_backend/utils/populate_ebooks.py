import sys
sys.path.append('../')

import os
import time
import threading
from openpyxl import load_workbook
from django.db import transaction

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'library_backend.settings')
import django
django.setup()

from databases.models import *
from concurrent.futures import ThreadPoolExecutor

def my_get_or_create(MyObj, **kwargs):
    try:
        obj = MyObj.objects.create(**kwargs)
    except InterruptedError:
        transaction.commit()
        obj = MyObj.objects.get(**kwargs)
    return obj

class EBookPopulate:
    def __init__(self):
        self.path = 'ebooks.xlsx'
        self.workbook = load_workbook(self.path, read_only=True)
        print("Workbook opened")
        self.sheet = self.workbook.active

    def ebook_populate(self, start):
        temp = start
        while temp < self.sheet.max_row:
            row = self.sheet[temp]
            subject = row[1].value.strip().title() if row[1].value else 'NA'
            subject = my_get_or_create(Subject, name=subject)
            
            author = row[2].value.strip().title() if row[2].value else 'NA'
            name = row[3].value.strip().title() if row[3].value else 'NA'
            
            publisher_name = row[5].value.strip().title() if row[5].value else 'NA'
            publisher = my_get_or_create(Publisher, name=publisher_name)
            
            url = row[8].value.strip() if row[8].value else 'NA'
            
            extra_data = {}
            extra_data['isbn'] = row[7].value if row[7].value else ''
            extra_data['edition'] = row[5].value.strip() if row[5].value else ''
            extra_data['year'] = row[6].value if row[6].value else ''
            
            ebook = my_get_or_create(EBook, name=name, author=author, publisher=publisher, subject=subject, url=url)
            ebook.extra_data = extra_data
            ebook.save()
            print(f"Added {name}")
            # move to the next 5th row
            temp += 5

def worker(start, ebook):
    print(f"Starting thread starting from {start}")
    ebook.populate(start)
    print(f"Thread finished starting from {start}")

if __name__ == "__main__":
    ebook = EBookPopulate()
    threads = []
    for i in range(1, 6):
        thread = threading.Thread(target=worker, args=(i, ebook))
        threads.append(thread)
        thread.start()
        #sleep for 5 seconds
        # time.sleep(5)
    for thread in threads:
        thread.join()
    print("Done")
